'''
Server-side script to update the sales and purchase list price 
of an item based on the purchase price and, in the case of
sales price, the margin uplift. At the same time, we take care
to insert new price_list items if they don't currently exist.
'''
import frappe
from frappe import _
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file as read_xlsx
from frappe.utils.xlsxutils import read_xls_file_from_attached_file as read_xls

import openpyxl as op

from datetime import date

from kanban_mods.utils.misc_functions import dump

@frappe.whitelist() 
def update_prices_from_quotation(doctype, item_code, new_price): 
    try:
        if item_code == "Shipping":
            # shipping is always variable for every order so we dont want 
            # to update the cost because we don't want it to pop up on
            # future quotations. 
            #
            return "Success"
        else:
            item = frappe.get_doc("Item", item_code) # the actual item itself
            new_price = float(new_price)
            update_buying_price(doctype, item_code, new_price, item) 
            sell_price = get_selling_price(new_price, item)
            update_selling_price(doctype, item_code, sell_price, item)
            return "Success"
    except Exception as ex:
        return ex
        
##### The  next two functions are the primary actions #######

def update_buying_price(doctype, item_code, price, item):

    buying_record = frappe.get_all(doctype, filters={'item_code': item_code, 'buying': True}) 
    if buying_record:
        frappe.db.set_value(doctype, buying_record[0].name, 'price_list_rate', price) 
    else: 
        price_list = item.brand+" Buying" 
        create_item_price(doctype, item_code, price_list, price, item) 
    frappe.db.commit()
    

def update_selling_price(doctype, item_code, price, item): 

    selling_record = frappe.get_all(doctype, filters={'item_code': item_code, 'selling': True})
    if selling_record:
        frappe.db.set_value(doctype, selling_record[0].name, 'price_list_rate', price) 
    else: 
        price_list = item.brand+" Selling" 
        create_item_price(doctype, item_code, price_list, price, item) 
    frappe.db.commit()  

##### And here are the secondary, supportingfunctions #######

def create_item_price(doctype, item_code, price_list, price, item):
    # Populate and save a new item_price record. The type of 
    # record is determined by the price list in use
    selling = price_list.endswith("Selling")
    buying  = not selling

    new_item_price = frappe.get_doc({'doctype':       doctype, 
                                    'item_code':      item_code,
                                    'item_name':      item.item_name, 
                                    'uom':            item.stock_uom,
                                    'brand':          item.brand,
                                    'currency':       'EUR',
                                    'selling':         selling,
                                    'buying':          buying,
                                    'price_list':      price_list, 
                                    'price_list_rate': price,
                                    'valid_from': date.today()
                                    }).insert()
    
####### This method is probably going to be redundant now ########
def get_selling_price(price, item):
    # # go fetch the correct rule ... you need the Buying rule data
    rule = frappe.get_doc("Pricing Rule", "Buying", {"item_group": item.item_group})

    margin = rule.margin_rate_or_amount
    margin_type = rule.margin_type
    # uplift the buying price to the sales "list price" and apply it
    # We always work with fixed percentages, but I added the redundant
    # case in case we ever need it
    if margin_type == "Percentage":
        new_s_p = price * (1 + margin/100)
    else:
        new_s_p= price + margin
    return new_s_p

@frappe.whitelist()
def bulk_insert():
    # The idea here is to create buying and selling Item Price records for every variant
    # This is just the driver for the actual action functions
    manufs = frappe.get_all("Manufacturer")
    filter = tuple(item.name for item in manufs)
    items = frappe.get_all("Item")
    ix = 0
    ix2 = 0
    for item in items:
        if item.name.endswith(filter):
            it = frappe.db.get_value("Item", item.name, ["item_name", "stock_uom", "brand", "item_group"], as_dict=1)
            if it.brand:
                update_buying_price("Item Price", it.item_name, 0, it)
                update_selling_price("Item Price", it.item_name, 0, it)
                ix += 1
                print(ix2," -- ",ix)
            else:
                raise Exception("Stock Item "+item.name+" does not have a brand. Please fix this asap")
        if ix == 100:
            ix = 0
            ix2 += 100
            frappe.db.commit()

    return "Its all ready for checking"

@frappe.whitelist()
def bulk_update_from_xls(file_name):
    types = ('XLS', 'XLSX')
    file_meta = frappe.get_value("File", {'file_name': file_name}, ['file_name', 'file_type'], as_dict=1)

    if file_meta and file_meta.file_type in types:
        # we have a valid file so lets process the data
        parsed_sheet = parse_workbook(file_name)
        dump(parsed_sheet)
        for row in parsed_sheet['pricelist']:
            # We are assuming the sheet has been normalised efore submission
            # So we don't need to mess around working out which column holds
            # which values - Item_name, Buying Price, Selling Price (optional)
            item_code = row[0]
            item = frappe.db.get_value("Item", item_code, ["item_name", "stock_uom", "brand", "item_group"], as_dict=1)
            bp = float(row[1])
            if len(row) > 2 and row[2] != None and row[2] != 0 and row[2] < bp: 
                sp = float(row[2])
            else:
                sp = float(get_selling_price(bp, item))

            print(item_code, bp, sp)
            update_buying_price("Item Price", item_code, bp, item)
            update_selling_price("Item Price", item_code, sp, item)
            
        return " go check the results ... "
    else:
        # return an error message
        return "Sorry but either " + file_name +" is not a valid Excel spreadsheet/workbook, or else it does not exist in the file system"
    
def parse_workbook(filename):
    _file = frappe.get_doc("File", {"file_name": filename})
    path_to_file = _file.get_full_path()
    book = op.load_workbook(filename=path_to_file, data_only=True, read_only=False)

    # Get the column titles
    column = 0
    col_titles = []
    current_sheet = book[book.sheetnames[0]]
    for col in current_sheet.iter_cols(1, current_sheet.max_column):
        column += 1
        col_titles.insert(column, col[0].value)

    # Get the actual data
    current_row = 0
    new_prices = []    
    current_row = 1    
    for sheet in book.sheetnames:
        current_sheet = book[sheet]
        # Iterate the loop to read the cell values
        for row in range(1, current_sheet.max_row):
            price_row = []
            column = 0
            for col in current_sheet.iter_cols(1, current_sheet.max_column):
                column += 1
                if column == 1:
                    price_row.insert(column, col[row].value+" - "+sheet)
                else:
                     price_row.insert(column, col[row].value)
            new_prices.insert(current_row, price_row)
            current_row += 1
    parsed_sheet = {}
    parsed_sheet['tites'] = col_titles
    parsed_sheet['pricelist'] = new_prices
    return parsed_sheet

        
    


        